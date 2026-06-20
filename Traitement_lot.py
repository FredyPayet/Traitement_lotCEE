import streamlit as st
import pandas as pd
import subprocess
import os
import re
import tempfile
import io
import zipfile
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

st.set_page_config(page_title="Traitement des lots CEE contrôles réglementaires", layout="wide")

st.title("📊 Traitement des lots CEE contrôles réglementaires")
st.markdown("Chargez votre fichier Excel (`.xls` ou `.xlsx`) pour générer automatiquement un fichier Excel par client.")

COL_LABELS = {
    "D": "Réf. interne",
    "E": "Nom du site",
    "F": "Adresse",
    "G": "Code postal",
    "H": "Ville",
    "I": "Raison sociale bénéficiaire",
    "BS": "Conclusion de l'audit",
    "BY": "Conformité après correction",
}

COLORS = {
    "satisfaisant":     "C6EFCE",
    "non_satisfaisant": "FFCCCC",
    "inaccessible":     "FFE0B2",
    "non_visite":       "E0E0E0",
    "header":           "2F5496",
}

NS_REF = {
    "EN-101": {
        "Surface": (
            "Il a été constaté un écart de plus de 10 % entre la surface déclarée et celle relevée sur site.",
            "Merci de nous confirmer la prise en compte de la surface mesurée, ou nous transmettre un document justificatif type plans pour contredire les mesures."
        ),
        "Ecart au feu": (
            "Il a été constaté une absence d'écart au feu entre l'isolant et une source de chaleur.",
            "Merci de faire réaliser un écart au feu de 10 cm, et de nous transmettre une attestation de reprise et des photos."
        ),
        "Epaisseur": (
            "Il a été constaté une épaisseur insuffisante pour atteindre la résistance thermique imposée par la fiche standardisée.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Homogénéité": (
            "Il a été constaté que la pose de l'isolant n'est pas homogène. La performance thermique est donc discontinue.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Risque électrique": (
            "Il a été constaté que la VMC est en contact avec l'isolant.",
            "Merci de faire réaliser un écart de 10 cm entre l'isolant et la VMC, et de nous transmettre une attestation de reprise et des photos."
        ),
        "Rehausse": (
            "Il a été constaté l'absence de rehausse de trappe.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Non réalisés": (
            "Il a été constaté que les travaux n'ont pas été réalisés.",
            "Merci de nous confirmer l'annulation de la valorisation pour cette opération."
        ),
        "Pare vapeur": (
            "Il a été constaté l'absence de pare vapeur, alors que l'élément est nécessaire.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Humidité": (
            "Il a été constaté des traces d'humidité sur les murs.",
            "Merci de faire installer un pare vapeur, et de nous transmettre une attestation de reprise et des photos."
        ),
        "Non éligible": (
            "Il a été constaté que la configuration de l'installation n'est pas éligible au CEE.",
            "Merci de nous confirmer l'annulation de la valorisation pour cette opération."
        ),
        "Trappe bloquée": (
            "Il a été constaté que la trappe est bloquée par l'isolant.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Déclaratif": (
            "Il a été constaté que les éléments décrits sur la facture ne correspondent pas à ce qui a été installé sur site.",
            "Merci de faire reprendre la facture et de nous la transmettre."
        ),
    },
    "EN-102": {
        "Surface": (
            "Il a été constaté un écart de plus de 10 % entre la surface déclarée et celle relevée sur site.",
            "Merci de nous confirmer la prise en compte de la surface mesurée, ou nous transmettre un document justificatif type plans pour contredire les mesures."
        ),
        "Epaisseur": (
            "Il a été constaté une épaisseur insuffisante pour atteindre la résistance thermique imposée par la fiche standardisée.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Protection intempérie": (
            "Il a été constaté que l'isolant n'est pas protégé des intempéries sur toute la surface.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Finition (ITI)": (
            "Il a été constaté que l'isolant n'est pas protégé des intempéries sur toute la surface.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Non réalisés": (
            "Il a été constaté que les travaux n'ont pas été réalisés.",
            "Merci de nous confirmer l'annulation de la valorisation pour cette opération."
        ),
        "Distance sol": (
            "Il a été constaté qu'il n'y a pas d'espace entre le départ de l'isolant et le sol, ce qui peut entraîner des remontées capillaires.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Qualité non manifeste": (
            "Il a été observé des non-qualités manifestes des travaux.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Déclaratif": (
            "Il a été constaté que les éléments décrits sur la facture ne correspondent pas à ce qui a été installé sur site.",
            "Merci de faire reprendre la facture et de nous la transmettre."
        ),
    },
    "EN-103": {
        "Surface": (
            "Il a été constaté un écart de plus de 10 % entre la surface déclarée et celle relevée sur site.",
            "Merci de nous confirmer la prise en compte de la surface mesurée, ou nous transmettre un document justificatif type plans pour contredire les mesures."
        ),
        "Ecart au feu": (
            "Il a été constaté une absence d'écart au feu entre l'isolant et une source de chaleur.",
            "Merci de faire réaliser un écart au feu de 10 cm, et de nous transmettre une attestation de reprise et des photos."
        ),
        "Epaisseur": (
            "Il a été constaté une épaisseur insuffisante pour atteindre la résistance thermique imposée par la fiche standardisée.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Homogénéité": (
            "Il a été constaté que la pose de l'isolant n'est pas homogène. La performance thermique est donc discontinue.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Risque électrique": (
            "Il a été constaté que des luminaires sont en contact avec l'isolant.",
            "Merci de faire réaliser un écart de 10 cm entre l'isolant et les luminaires, et de nous transmettre une attestation de reprise et des photos."
        ),
        "Non réalisés": (
            "Il a été constaté que les travaux n'ont pas été réalisés.",
            "Merci de nous confirmer l'annulation de la valorisation pour cette opération."
        ),
        "Humidité": (
            "Il a été constaté des traces d'humidité sur les murs.",
            "Merci de faire installer un pare vapeur, et de nous transmettre une attestation de reprise et des photos."
        ),
        "Fixation": (
            "Il a été constaté un défaut de fixation sur les panneaux d'isolant.",
            "Merci de faire reprendre les travaux et de nous transmettre une attestation de reprise et des photos."
        ),
        "Morcellement (points particuliers non isolés)": (
            "Il a été constaté qu'une poutre n'est pas isolée et entraîne une discontinuité d'isolant.",
            "Merci de faire isoler la poutre pour enlever le pont thermique, et de nous transmettre une attestation de reprise et des photos. L'isolant peut être de fine couche."
        ),
        "Non éligible": (
            "Il a été constaté que la configuration de l'installation n'est pas éligible au CEE.",
            "Merci de nous confirmer l'annulation de la valorisation pour cette opération."
        ),
        "Déclaratif": (
            "Il a été constaté que les éléments décrits sur la facture ne correspondent pas à ce qui a été installé sur site.",
            "Merci de faire reprendre la facture et de nous la transmettre."
        ),
    },
    "EN-105": {
        "Surface": ("", ""),
        "Epaisseur": ("", ""),
        "Homogénéité": ("", ""),
        "Etanchéité": ("", ""),
        "Ecart au feu": ("", ""),
        "Non réalisés": ("", ""),
        "Document": ("", ""),
    },
}

FICHE_COLS = {
    "EN-101": {"extra": ["BS", "BY"], "labels": {"BS": "Résultat du contrôle sur site", "BY": "Résultat du contrôle par contact"}},
    "EN-102": {"extra": ["BS", "BY"], "labels": {"BS": "Résultat du contrôle sur site", "BY": "Résultat du contrôle par contact"}},
    "EN-103": {"extra": ["BS", "BY"], "labels": {"BS": "Résultat du contrôle sur site", "BY": "Résultat du contrôle par contact"}},
    "TH-106": {"extra": ["AI"], "labels": {"AI": "Résultat du contrôle par contact"}},
    "TH-158": {"extra": ["AI"], "labels": {"AI": "Résultat du contrôle par contact"}},
    "TH-127": {"extra": ["BH", "BN"], "labels": {"BH": "Résultat du contrôle sur site", "BN": "Résultat du contrôle par contact"}},
    "EN-105": {"extra": ["BH", "BN"], "labels": {"BH": "Résultat du contrôle sur site", "BN": "Résultat du contrôle par contact"}},
    "TH-107": {"extra": ["AL"], "labels": {"AL": "Résultat du contrôle par contact"}},
    "TH-107-SE": {"extra": ["AL"], "labels": {"AL": "Résultat du contrôle par contact"}},
    "EN-104": {"extra": ["AL"], "labels": {"AL": "Résultat du contrôle par contact"}},
    "TH-112": {"extra": ["AK"], "labels": {"AK": "Résultat du contrôle par contact"}},
}

# Messages dossier incomplet
MESSAGES_INCOMPLET = {
    "AH non reçue": (
        "Nous n'avons pas encore reçu l'attestation sur l'honneur pour les fiches complètes du dossier. "
        "Vous les trouverez ci-joint. Merci de nous les retourner avec les parties B, C et BS datées, "
        "signées et tamponnées. Le document doit nous être transmis en version numérique et originale par voie postale."
    ),
    "Documents non conformes": (
        "Un document administratif est non conforme. Merci de vous rendre sur le dossier "
        "et nous fournir les éléments demandés."
    ),
}


def get_fiche_extra_cols(fiche: str) -> tuple[list, dict]:
    cfg = FICHE_COLS.get(fiche, {"extra": ["BS", "BY"], "labels": {"BS": "Résultat du contrôle sur site", "BY": "Résultat du contrôle par contact"}})
    return cfg["extra"], cfg["labels"]


def get_col_by_label(fiche: str, label: str) -> str | None:
    extra_cols, extra_labels = get_fiche_extra_cols(fiche)
    for col, lbl in extra_labels.items():
        if lbl == label:
            return col
    return None


def get_seuils(fiche: str, annee: int) -> dict:
    has_site    = get_col_by_label(fiche, "Résultat du contrôle sur site")    is not None
    has_contact = get_col_by_label(fiche, "Résultat du contrôle par contact") is not None

    if fiche == "EN-101" and annee in (2019, 2020, 2021):
        return {"seuil_s_site": 10.0}
    if fiche == "EN-103" and annee in (2019, 2020, 2021):
        return {"seuil_s_site": 20.0}
    if fiche == "EN-102" and annee == 2021:
        return {"seuil_s_site": 10.0, "seuil_s_contact": 20.0}

    if has_site and has_contact:
        table = {2022: (7.5, 15.0), 2023: (10.0, 20.0), 2024: (12.5, 25.0)}
        s, c = table.get(annee, (15.0, 30.0))
        return {"seuil_s_site": s, "seuil_s_contact": c}

    if has_contact and not has_site:
        table_c = {2022: 15.0, 2023: 20.0, 2024: 25.0}
        c = table_c.get(annee, 30.0)
        return {"seuil_s_contact": c}

    if has_site and not has_contact:
        table_s = {2022: 7.5, 2023: 10.0, 2024: 12.5}
        s = table_s.get(annee, 15.0)
        return {"seuil_s_site": s}

    return {}


def compute_taux(data: pd.DataFrame, fiche: str, total_ops: int) -> dict:
    result = {}
    col_site    = get_col_by_label(fiche, "Résultat du contrôle sur site")
    col_contact = get_col_by_label(fiche, "Résultat du contrôle par contact")

    annee = None
    if "Q" in data.columns:
        dates_valides = data["Q"].dropna()
        if not dates_valides.empty:
            annee = dates_valides.max().year
    result["annee"] = annee
    result["seuils"] = get_seuils(fiche, annee) if annee else {}

    if col_site and col_site in data.columns:
        vals_site = data[col_site].astype(str).str.lower().str.strip()
        nb_s_site         = int((vals_site == "satisfaisant").sum())
        nb_ns_site        = int((vals_site == "non satisfaisant").sum())
        nb_controles_site = int((vals_site != "non visité").sum())

        result["taux_s_site"]       = nb_s_site  / total_ops * 100 if total_ops > 0 else 0
        result["taux_ns_site"]      = nb_ns_site / nb_controles_site * 100 if nb_controles_site > 0 else 0
        result["nb_s_site"]         = nb_s_site
        result["nb_ns_site"]        = nb_ns_site
        result["nb_controles_site"] = nb_controles_site
        result["total_ops"]         = total_ops

    if col_contact and col_contact in data.columns:
        vals_contact = data[col_contact].astype(str).str.lower().str.strip()
        nb_s_contact = int((vals_contact == "satisfaisant").sum())
        result["taux_s_contact"] = nb_s_contact / total_ops * 100 if total_ops > 0 else 0
        result["nb_s_contact"]   = nb_s_contact

    return result


def col_letter_to_idx(letter: str) -> int:
    idx = 0
    for ch in letter.upper().strip():
        idx = idx * 26 + (ord(ch) - ord('A') + 1)
    return idx - 1


def convert_xls_to_xlsx(xls_path: str) -> str:
    out_dir = tempfile.mkdtemp()
    result = subprocess.run(
        ["libreoffice", "--headless", "--convert-to", "xlsx", "--outdir", out_dir, xls_path],
        capture_output=True, text=True
    )
    if result.returncode != 0:
        raise RuntimeError(f"Conversion LibreOffice échouée :\n{result.stderr}")
    base = os.path.splitext(os.path.basename(xls_path))[0]
    return os.path.join(out_dir, base + ".xlsx")


def load_dataframe(file_bytes: bytes, filename: str) -> pd.DataFrame:
    ext = os.path.splitext(filename)[1].lower()
    if ext in (".xlsx", ".xlsm"):
        return pd.read_excel(io.BytesIO(file_bytes), header=None)
    with tempfile.NamedTemporaryFile(suffix=".xls", delete=False) as tmp:
        tmp.write(file_bytes)
        tmp_path = tmp.name
    try:
        xlsx_path = convert_xls_to_xlsx(tmp_path)
        return pd.read_excel(xlsx_path, header=None)
    finally:
        os.unlink(tmp_path)


def extract_tables(df: pd.DataFrame, fiche: str = "EN-101"):
    header_idx = None
    for i, row in df.iterrows():
        for val in row:
            if isinstance(val, str) and "REFERENCE EMMY" in val.upper():
                header_idx = i
                break
        if header_idx is not None:
            break
    if header_idx is None:
        raise ValueError("Impossible de trouver la ligne d'en-tête (cherche 'REFERENCE EMMY').")

    extra_cols, _ = get_fiche_extra_cols(fiche)
    base_cols = ["D", "E", "F", "G", "H", "I"]
    all_col_letters = base_cols + extra_cols + ["N", "O", "Q"]
    needed_cols = [col_letter_to_idx(l) for l in all_col_letters]

    if max(needed_cols) >= df.shape[1]:
        raise ValueError(f"Le fichier n'a que {df.shape[1]} colonnes (colonne {all_col_letters[-1]} attendue).")

    data_raw = df.iloc[header_idx + 1:, needed_cols].copy()
    data_raw.columns = all_col_letters
    data_raw = data_raw.dropna(subset=["I"], how="all")
    data_raw = data_raw[data_raw["I"].astype(str).str.strip() != ""]

    if data_raw.empty:
        raise ValueError("Aucune ligne de données trouvée. Vérifiez que la colonne I est remplie.")

    for col in extra_cols:
        data_raw[col] = data_raw[col].fillna("non visité")
        data_raw[col] = data_raw[col].apply(lambda v: "non visité" if str(v).strip() == "" else v)

    for col in ["N", "O"]:
        data_raw[col] = pd.to_numeric(data_raw[col], errors="coerce").fillna(0)

    data_raw["Q"] = pd.to_datetime(data_raw["Q"], errors="coerce")

    clients = sorted(data_raw["I"].dropna().unique().tolist())
    return data_raw.reset_index(drop=True), clients


def sanitize_filename(name: str) -> str:
    return re.sub(r'[\\/*?:"<>|]', "", name).strip()


def get_cell_color(value: str):
    v = str(value).lower()
    if "non satisfaisant" in v:
        return COLORS["non_satisfaisant"]
    elif "satisfaisant" in v:
        return COLORS["satisfaisant"]
    elif "inaccessible" in v or "non vérifiable" in v:
        return COLORS["inaccessible"]
    elif "non visité" in v:
        return COLORS["non_visite"]
    return None


def build_client_excel(df_client: pd.DataFrame, client_name: str, lot_destination: str = "", fiche: str = "EN-101") -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = client_name[:31]

    extra_cols, extra_labels = get_fiche_extra_cols(fiche)
    base_keys = ["D", "E", "F", "G", "H", "I"]
    col_keys = base_keys + extra_cols
    all_labels = {**COL_LABELS, **extra_labels}
    headers = [all_labels.get(k, k) for k in col_keys]

    if lot_destination:
        col_keys_ext = col_keys + ["LOT_DEST"]
        headers_ext = headers + ["Nouveau lot de contrôle"]
    else:
        col_keys_ext = col_keys
        headers_ext = headers

    colored_col_indices = [col_keys_ext.index(c) + 1 for c in extra_cols if c in col_keys_ext]

    thin = Side(style="thin", color="000000")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    header_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    header_fill = PatternFill("solid", fgColor=COLORS["header"])
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_align = Alignment(vertical="center", wrap_text=True)

    ws.row_dimensions[1].height = 40
    for col_num, header in enumerate(headers_ext, 1):
        cell = ws.cell(row=1, column=col_num, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = border

    for row_num, (_, row) in enumerate(df_client.iterrows(), 2):
        ws.row_dimensions[row_num].height = 20
        for col_num, key in enumerate(col_keys_ext, 1):
            val = lot_destination if key == "LOT_DEST" else row[key]
            cell = ws.cell(row=row_num, column=col_num, value=val)
            cell.alignment = cell_align
            cell.border = border
            cell.font = Font(name="Arial", size=10)
            if col_num in colored_col_indices:
                color = get_cell_color(str(val))
                if color:
                    cell.fill = PatternFill("solid", fgColor=color)

    base_widths = [20, 30, 35, 12, 20, 35]
    extra_widths = [30] * len(extra_cols)
    col_widths = base_widths + extra_widths + ([25] if lot_destination else [])
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def build_filename(client_name: str, num_dossier: str, num_lot: str) -> str:
    name = sanitize_filename(client_name)[:50]
    dossier = sanitize_filename(num_dossier.strip()) if num_dossier.strip() else "sans_dossier"
    lot = sanitize_filename(num_lot.strip()) if num_lot.strip() else "lot_non_renseigné"
    return f"{name}_{dossier}_{lot}.xlsx"


def get_ns_adresses(df_client: pd.DataFrame, conclusion_col: str = "BS") -> list:
    if conclusion_col not in df_client.columns:
        conclusion_col = df_client.columns[6] if len(df_client.columns) > 6 else "BS"
    ns_rows = df_client[df_client[conclusion_col].astype(str).str.lower().str.contains("non satisfaisant", na=False)]
    result = []
    for _, row in ns_rows.iterrows():
        adresse = " ".join(filter(None, [
            str(row["F"]).strip() if pd.notna(row["F"]) else "",
            str(row["G"]).strip() if pd.notna(row["G"]) else "",
            str(row["H"]).strip() if pd.notna(row["H"]) else "",
        ]))
        ref_d = str(row["D"]).strip() if pd.notna(row["D"]) else ""
        fiche = detect_fiche(ref_d)
        if adresse.strip():
            result.append({"adresse": adresse, "fiche": fiche})
    return result


def detect_fiche(ref: str) -> str | None:
    m = re.search(r"(EN|BAR|BAT|RES|IND|AGR|TRA)-?\d{3}", ref, re.IGNORECASE)
    if m:
        return m.group(0).upper()
    m2 = re.search(r"[A-Z]{2,3}-\d{3}", ref, re.IGNORECASE)
    return m2.group(0).upper() if m2 else None


def build_message(taux_choix: str, lot_label: str, ns_adresses_causes: list,
                  lot_destination: str = "", delais_courts: bool = False,
                  tous_non_visites: bool = False,
                  incomplet_choix: list = None) -> str:

    bloc_ns = ""
    if ns_adresses_causes:
        lignes = []
        for adresse, nc_items in ns_adresses_causes:
            lignes.append(f"• {adresse} :")
            for nc_label, msg_complet in nc_items:
                if msg_complet.strip():
                    lignes.append(f"\t• {msg_complet.strip()}")
        bloc_ns = "\n\n" + "\n".join(lignes)

    # Bloc dossier incomplet (avant "Vous trouverez ci-joint...")
    bloc_incomplet = ""
    if incomplet_choix:
        if len(incomplet_choix) == 1:
            bloc_incomplet = f"\n\n{MESSAGES_INCOMPLET[incomplet_choix[0]]}"
        else:
            lignes_inc = [f"• {MESSAGES_INCOMPLET[c]}" for c in incomplet_choix]
            bloc_incomplet = "\n\n" + "\n".join(lignes_inc)

    bloc_destination = ""
    if lot_destination.strip():
        bloc_destination = f"\n\nLes opérations concernées devront être représentées dans le lot de destination suivant : {lot_destination.strip()}."

    bloc_delais = ""
    if delais_courts:
        bloc_delais = "\n\nLa date de fin de travaux est inférieure à 3 mois. Nous ne pouvons plus intégrer le dossier dans un nouveau lot de contrôle pour validation. Merci de nous fournir un document du marché permettant de repousser cette date, ou de nous confirmer l'annulation du dossier."

    fin = f"{bloc_destination}{bloc_delais}\n\nCordialement,"
    odicee = "" if tous_non_visites else "\n\nLes rapports de contrôle sont disponibles sur ODICEE, si vos opérations ont été contrôlées."

    corps = {
        "Taux OK": (
            f"Bonjour,\n\n\n"
            f"Pour votre information, nous avons reçu le retour du lot de contrôle {lot_label}.\n\n"
            f"Tous les taux réglementaires sont respectés. Toutes les opérations du lot relatif "
            f"à la fiche travaux peuvent être finalisées."
            f"{bloc_incomplet}\n\n"
            f"Vous trouverez ci-joint les résultats des contrôles pour vos opérations."
            f"{bloc_ns}{odicee}\n\n{fin}"
        ),
        "Taux NS KO": (
            f"Bonjour,\n\n\n"
            f"Pour votre information, nous avons reçu le retour du lot de contrôle {lot_label}.\n\n"
            f"Le taux d'opérations contrôlées non satisfaisantes dépasse les 10 %. "
            f"Nous ne pouvons finaliser que les opérations qui ont été contrôlées. "
            f"Les opérations non visitées doivent être représentées dans un nouveau lot."
            f"{bloc_incomplet}\n\n"
            f"Vous trouverez ci-joint les résultats des contrôles pour vos opérations."
            f"{bloc_ns}{odicee}\n\n{fin}"
        ),
        "Tous taux KO": (
            f"Bonjour,\n\n\n"
            f"Pour votre information, nous avons reçu le retour du lot de contrôle {lot_label}.\n\n"
            f"Les taux réglementaires ne sont pas atteints. Nous ne pouvons finaliser aucune opération "
            f"dans ce lot. Les opérations doivent être représentées dans un nouveau lot."
            f"{bloc_incomplet}\n\n"
            f"Vous trouverez ci-joint les résultats des contrôles pour vos opérations."
            f"{bloc_ns}{odicee}\n\n{fin}"
        ),
    }
    return corps[taux_choix]


def copy_and_download_button(text: str, xlsx_bytes: bytes, filename: str, client_name: str, key: str):
    escaped = text.replace("\\", "\\\\").replace("`", "\\`").replace("$", "\\$")
    label   = f"📋 Copier le message — {client_name}"

    st.components.v1.html(
        f"""
        <button onclick="
            navigator.clipboard.writeText(`{escaped}`)
            .then(() => {{
                this.innerText = '✅ Message copié !';
                setTimeout(() => this.innerText = '{label}', 2500);
            }})
            .catch(() => {{
                this.innerText = '⚠️ Utilisez Ctrl+C';
                setTimeout(() => this.innerText = '{label}', 2500);
            }});
        "
        style="
            background-color:#2F5496; color:white; border:none;
            padding:10px 20px; border-radius:6px; font-size:14px;
            cursor:pointer; font-family:Arial,sans-serif; width:100%;
        ">{label}</button>
        """,
        height=55,
    )

    st.download_button(
        label=f"⬇️ Télécharger les résultats Excel — {client_name}",
        data=xlsx_bytes,
        file_name=filename,
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        key=f"dl_{key}",
        use_container_width=True,
    )


def afficher_taux(taux: dict, fiche: str):
    has_site    = "taux_s_site"    in taux
    has_contact = "taux_s_contact" in taux
    seuils      = taux.get("seuils", {})
    annee       = taux.get("annee")

    if annee:
        st.caption(f"📅 Date d'engagement la plus récente : **{annee}** — seuils appliqués : "
                   + (f"S site ≥ {seuils.get('seuil_s_site', '—')} %" if 'seuil_s_site' in seuils else "")
                   + ("  |  " if 'seuil_s_site' in seuils and 'seuil_s_contact' in seuils else "")
                   + (f"S contact ≥ {seuils.get('seuil_s_contact', '—')} %" if 'seuil_s_contact' in seuils else ""))

    nb_cartes = sum([has_site * 2, has_contact])
    cols = st.columns(nb_cartes if nb_cartes > 0 else 1)
    col_idx = 0

    if has_site:
        seuil_s  = seuils.get("seuil_s_site", 0)
        seuil_ns = 10.0
        ok_s  = taux["taux_s_site"]  >= seuil_s
        ok_ns = taux["taux_ns_site"] <= seuil_ns
        total = taux.get("total_ops", taux["nb_s_site"] + taux["nb_ns_site"])

        with cols[col_idx]:
            color = "#C6EFCE" if ok_s else "#FFCCCC"
            icone = "✅" if ok_s else "❌"
            st.markdown(
                f"<div style='background:{color};padding:12px;border-radius:8px;text-align:center'>"
                f"<b>Taux S sur site</b> {icone}<br>"
                f"<span style='font-size:22px;font-weight:bold'>{taux['taux_s_site']:.1f} %</span><br>"
                f"<small>{taux['nb_s_site']} S / {total} opérations — seuil ≥ {seuil_s} %</small>"
                f"</div>",
                unsafe_allow_html=True,
            )
        col_idx += 1

        with cols[col_idx]:
            color = "#C6EFCE" if ok_ns else "#FFCCCC"
            icone = "✅" if ok_ns else "❌"
            st.markdown(
                f"<div style='background:{color};padding:12px;border-radius:8px;text-align:center'>"
                f"<b>Taux NS sur site</b> {icone}<br>"
                f"<span style='font-size:22px;font-weight:bold'>{taux['taux_ns_site']:.1f} %</span><br>"
                f"<small>{taux['nb_ns_site']} NS / {taux['nb_controles_site']} contrôlées — seuil ≤ {seuil_ns} %</small>"
                f"</div>",
                unsafe_allow_html=True,
            )
        col_idx += 1

    if has_contact:
        seuil_c = seuils.get("seuil_s_contact", 0)
        ok_c    = taux["taux_s_contact"] >= seuil_c
        total   = taux.get("total_ops", 0)
        icone   = "✅" if ok_c else "❌"
        with cols[col_idx]:
            color = "#C6EFCE" if ok_c else "#FFCCCC"
            st.markdown(
                f"<div style='background:{color};padding:12px;border-radius:8px;text-align:center'>"
                f"<b>Taux S par contact</b> {icone}<br>"
                f"<span style='font-size:22px;font-weight:bold'>{taux['taux_s_contact']:.1f} %</span><br>"
                f"<small>{taux['nb_s_contact']} S / {total} opérations — seuil ≥ {seuil_c} %</small>"
                f"</div>",
                unsafe_allow_html=True,
            )


# ─── UI ─────────────────────────────────────────────────────────────────────

st.subheader("📋 Informations du lot")

col1, col2, col3 = st.columns(3)
with col1:
    num_lot = st.text_input("Numéro de lot", placeholder="ex : LOT-2024-001")
with col2:
    taux_choix = st.selectbox(
        "Résultat du lot",
        options=["Taux OK", "Taux NS KO", "Tous taux KO"],
    )
with col3:
    fiches_dispo = sorted(FICHE_COLS.keys())
    fiche_globale = st.selectbox(
        "Fiche CEE du lot",
        options=fiches_dispo,
        key="fiche_globale",
    )

st.markdown("---")

uploaded = st.file_uploader(
    "Choisir un fichier Excel (.xls / .xlsx)",
    type=["xls", "xlsx", "xlsm"],
)

if uploaded:
    with st.spinner("Lecture du fichier…"):
        try:
            file_bytes = uploaded.read()
            df_raw = load_dataframe(file_bytes, uploaded.name)
            data, clients = extract_tables(df_raw, fiche_globale)
        except Exception as e:
            st.error(f"❌ Erreur lors de la lecture : {e}")
            st.stop()

    st.success(f"✅ Fichier chargé — **{len(data)}** ligne(s) · **{len(clients)}** client(s) détecté(s)")

    total_ops = int(data["D"].notna().sum() - (data["D"].astype(str).str.strip() == "").sum())
    taux_lot = compute_taux(data, fiche_globale, total_ops)
    if taux_lot:
        st.markdown(f"#### 📊 Taux réglementaires du lot — **{total_ops} opération(s)**")
        afficher_taux(taux_lot, fiche_globale)

    st.markdown("---")

    lot_label = num_lot.strip() if num_lot.strip() else "[numéro de lot non renseigné]"

    couleur_bandeau = {
        "Taux OK":      "#e8f5e9",
        "Taux NS KO":   "#fff3e0",
        "Tous taux KO": "#ffebee",
    }[taux_choix]

    st.sidebar.header("Filtres & Export")
    selected = st.sidebar.multiselect("Clients à afficher", options=clients, default=clients)

    if st.sidebar.button("⬇️ Télécharger tous les fichiers (ZIP)"):
        zip_buf = io.BytesIO()
        with zipfile.ZipFile(zip_buf, "w", zipfile.ZIP_DEFLATED) as zf:
            for client in clients:
                df_c = data[data["I"] == client].copy()
                dest_active_key = f"dest_active_{client}"
                dest_key = f"lot_dest_{client}"
                prefixes_c = []
                for val in df_c["D"].dropna():
                    digits_c = re.sub(r'\D', '', str(val))[:6]
                    if digits_c and digits_c not in prefixes_c:
                        prefixes_c.append(digits_c)
                num_dossier_c = " / ".join(prefixes_c) if prefixes_c else ""
                lot_dest_c = st.session_state.get(dest_key, "") if st.session_state.get(dest_active_key, False) else ""
                xlsx_bytes = build_client_excel(df_c, client, lot_dest_c, fiche_globale)
                fname = build_filename(client, num_dossier_c, num_lot)
                zf.writestr(fname, xlsx_bytes)
        st.sidebar.download_button(
            label="📦 Télécharger le ZIP",
            data=zip_buf.getvalue(),
            file_name="tableaux_clients.zip",
            mime="application/zip",
        )

    if not selected:
        st.info("Aucun client sélectionné dans le panneau de gauche.")
    else:
        for client in selected:
            df_client = data[data["I"] == client].copy().reset_index(drop=True)

            volume = (pd.to_numeric(df_client["N"], errors="coerce").fillna(0) +
                      pd.to_numeric(df_client["O"], errors="coerce").fillna(0)).sum() * 0.001
            volume_str = f"{volume:,.3f} MWhc".replace(",", " ")
            volume_bold = volume >= 2000
            label_volume = f"**{volume_str}**" if volume_bold else volume_str

            with st.expander(f"🏢 {client}  ({len(df_client)} opération(s)) — {label_volume}", expanded=True):

                prefixes = []
                for val in df_client["D"].dropna():
                    digits = re.sub(r'\D', '', str(val))[:6]
                    if digits and digits not in prefixes:
                        prefixes.append(digits)
                num_dossier = " / ".join(prefixes) if prefixes else ""
                st.info(f"📁 Numéro(s) de dossier : **{num_dossier}**" if num_dossier else "📁 Aucun numéro de dossier détecté")

                # ── Options par client ───────────────────────────────────────
                dest_active_key = f"dest_active_{client}"
                dest_key        = f"lot_dest_{client}"
                delais_key      = f"delais_{client}"
                incomplet_key   = f"incomplet_{client}"
                incomplet_choix_key = f"incomplet_choix_{client}"

                col_opt1, col_opt2, col_opt3 = st.columns(3)
                with col_opt1:
                    activer_destination = st.checkbox("Spécifier un lot de destination", key=dest_active_key)
                with col_opt2:
                    delais_courts = st.checkbox("⏳ Délai restant pour contrôle < 3 mois", key=delais_key)
                with col_opt3:
                    dossier_incomplet = st.checkbox("📂 Dossier incomplet", key=incomplet_key)

                lot_destination = ""
                if activer_destination:
                    lot_destination = st.text_input(
                        "Lot de destination",
                        placeholder="ex : LOT-2024-002",
                        key=dest_key,
                    )

                # Menu déroulant dossier incomplet (sur la même ligne que la coche)
                incomplet_choix = []
                if dossier_incomplet:
                    incomplet_choix = st.multiselect(
                        "Type d'incomplet",
                        options=list(MESSAGES_INCOMPLET.keys()),
                        default=[],
                        key=incomplet_choix_key,
                        label_visibility="collapsed",
                    )

                filename = build_filename(client, num_dossier, num_lot)

                st.markdown("---")

                extra_cols_fiche, _ = get_fiche_extra_cols(fiche_globale)
                conclusion_col = extra_cols_fiche[0]
                ns_adresses = get_ns_adresses(df_client, conclusion_col)
                tous_non_visites = df_client[conclusion_col].astype(str).str.lower().str.strip().eq("non visité").all()
                ns_adresses_causes = []

                if ns_adresses:
                    st.markdown("**⚠️ Opérations non satisfaisantes — saisir les non-conformités :**")

                    for i, item in enumerate(ns_adresses):
                        adresse = item["adresse"]
                        st.markdown(f"📍 **{adresse}**")

                        fiche_sel = fiche_globale
                        nc_list = list(NS_REF.get(fiche_sel, {}).keys())

                        count_key = f"nc_count_{client}_{i}"
                        if count_key not in st.session_state:
                            st.session_state[count_key] = 1

                        nc_items = []
                        for j in range(st.session_state[count_key]):
                            col_sel, col_btn = st.columns([10, 1])
                            with col_sel:
                                nc_sel = st.selectbox(
                                    f"Non-conformité {j+1}",
                                    options=["— Sélectionner —"] + nc_list,
                                    key=f"nc_sel_{client}_{i}_{j}",
                                    label_visibility="collapsed",
                                )
                            if j == st.session_state[count_key] - 1:
                                with col_btn:
                                    if st.button("➕", key=f"add_{client}_{i}_{j}", help="Ajouter une non-conformité"):
                                        st.session_state[count_key] += 1
                                        st.rerun()

                            if nc_sel and nc_sel != "— Sélectionner —" and nc_sel in NS_REF.get(fiche_sel, {}):
                                msg, act = NS_REF[fiche_sel][nc_sel]
                                msg_complet = f"{msg} {act}".strip()
                                nc_items.append((nc_sel, msg_complet))

                        ns_adresses_causes.append((adresse, nc_items))
                        st.markdown("")

                    st.markdown("---")

                message = build_message(
                    taux_choix, lot_label, ns_adresses_causes,
                    lot_destination, delais_courts, tous_non_visites,
                    incomplet_choix if dossier_incomplet else [],
                )

                st.markdown("**✉️ Message à envoyer au client :**")
                st.markdown(
                    f"<div style='background-color:{couleur_bandeau}; padding:16px; "
                    f"border-radius:8px; white-space:pre-wrap; font-family:Arial; font-size:14px;'>"
                    f"{message}</div>",
                    unsafe_allow_html=True,
                )
                st.markdown("---")
                xlsx_bytes = build_client_excel(df_client, client, lot_destination, fiche_globale)
                copy_and_download_button(message, xlsx_bytes, filename, client, key=f"copy_{client}")